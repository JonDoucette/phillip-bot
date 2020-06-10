import discord
import json
import random
import datetime
import openpyxl

from discord.ext import commands, tasks
from itertools import cycle

client = discord.Client()
client = commands.Bot(command_prefix = '!', case_insensitive = True)
client.remove_command('help') #Removes the old command

numbers = []
used_numbers = []
reactId = 0

tilt = {}


@client.event
async def on_ready():
	await client.change_presence(status = discord.Status.online, activity = discord.Game('Executing Muzz'))
	print('Bot is ready')

#Method for obtaining random quote from JSON file
def randomQuote():

	f = open('json.json')

	data = json.load(f)

	response = ''
	number = random.randint(0, len(data['messages'])) #Generates a random number to the length of the quotes
	response = str(data['messages'][number]['content']) #Finds the quote
	embed = discord.Embed(color = 0x607d8b, description = response)

	while len(response) < 8: #If the response is a short quote of less than 8 characters, try again

		if len(response) == 0: #If the response is blank (i.e. there is only an image)
			response = str(data['messages'][number]['attachments'][0]['url']) #Sets response to the URL 
			embed.set_image(url = response) #Adds an image to the embed
			

		else:
			#Generates a new quote 
			number = random.randint(0, len(data['messages']))
			response = str(data['messages'][number]['content'])

			#Creates new embed with updated description
			embed2 = discord.Embed(color = 0x607d8b, description = response)
			embed = embed2
			

	
	return (embed)


@client.command()
async def quote(ctx):
	response = randomQuote()
	await ctx.send(embed = response)

@client.command()
async def update(ctx):
	embed = discord.Embed(color = 0x607d8b, description = 'Last updated on 06-09-2020')
	await ctx.send(embed = embed)

@client.command()
async def creator(ctx):
	embed = discord.Embed(color = 0x607d8b, description = 'Created by Jon Doucette')
	await ctx.send(embed = embed)

@client.command()
async def idiot(ctx):
	await ctx.send('No you')

@client.command(aliases = ['agent'])
async def agents(ctx, *, users):
	agents = ['Sage', 'Brimstone', 'Breach', 'Phoenix', 'Reyna',
			  'Omen', 'Sova', 'Jett', 'Viper', 'Cypher']	

	users = users.split()
	response = ''

	for i, user in enumerate(users):
		number = random.randint(0, len(agents)-1)
		response += (str(i+1) + '. '+ user + ' - ' + agents[number] + '\n')
		del agents[number]

	embed = discord.Embed(color = 0xFF0000, timestamp = datetime.datetime.utcnow(),
		description = response, title = 'Valorant Picks')

	#Sets who created the agents
	author = ctx.message.author
	if author.nick == None:
		author = author.display_name
	else:
		author = author.nick
	embed.set_author(name = 'Agent selection created by ' + author)


	await ctx.send(embed = embed)

#@client.command(aliases = ['tilt'])
#async def tilted(ctx, user):

#	print(user)

#	wb = openpyxl.load_workbook('tilted.xlsx')
#	sheet = wb['output']
#	for rowNum in range(1, sheet.max_row+1):
#		excelName = sheet.cell(row=rowNum, column=1).value
#		if excelName == user:
#			previous = sheet.cell(row=rowNum, column=2).value
#			tiltAmount = random.randint(0,200)
#			new = previous - tiltAmount
#			sheet.cell(row=rowNum, column=2).value = new

#	wb.save('tilted.xlsx')
#	wb.close()

#	await ctx.send(user + ' was tilted by ' + str(tiltAmount) + '. Their new mmr is: ' + str(new))

	
@client.command()
async def help(ctx):

	#Sends these privately to the user
	#author = ctx.message.author
	#await author.send(embed = embed)
	embed = discord.Embed(
		colour = discord.Colour.orange(),
		title = 'Help'
		)

	fields = [('`!quote`', 'Pulls a quote from the Mind-Break-Book', False),
			  ('`!update`', 'Last time the bot was updated', False),
			  ('`!game` or `!games` or `!movies`', "Enter games after command to create a poll\n Place a space between each game like:\n `!game Valorant League TTT`", False),
			  ('`!creator`', 'Lists the creator of the bot', False),
			  ('`!agents`', 'Picks Valorant Agents for the user to play \n Place a space between each user like: \n`!agents @user1 @user2` ', False)]

	for name, value, inline in fields:
		embed.add_field(name=name, value=value, inline = inline)


	await ctx.send(embed = embed)

@client.command(aliases = ['games','movies','movie'])
async def game(ctx, *, games):

	#await ctx.channel.purge(limit=1) #Will remove the command message 

	global numbers
	global used_numbers
	global reactId

	games = games.split()
	values = ''
	i = 1 #Starts the Value at A
	numbers = ['1️⃣','2️⃣','3️⃣','4️⃣','5️⃣','6️⃣','7️⃣','8️⃣','9️⃣','🔟'] #Reactions

	for game in games:
		values += ( str(i) + '. ' + game + '\n')
		i += 1

	embed = discord.Embed(color = 0xFF0000, timestamp = datetime.datetime.utcnow(), 
		title = 'Vote on an Option', description = values)

	#Sets who created the poll
	author = ctx.message.author
	if author.nick == None:
		author = author.display_name
	else:
		author = author.nick
	embed.set_author(name = 'Poll created by ' + author)


	#embed.add_field(name = 'Games', value = values, inline = False)
	reacted_message = await ctx.send(embed = embed)

	reactId = reacted_message.id

	#Adds reactions found in Numbers list based on games
	for num in range(len(games)):
		await reacted_message.add_reaction(numbers[num])
		used_numbers.append(numbers[num])



@client.event
async def on_reaction_add(reaction, user):
	global reactId
	global used_numbers
	if reaction.message.id == reactId:
		if reaction.emoji not in used_numbers:
			await reaction.clear()


#@tasks.loop(seconds = 5)
#async def remove_reactions():
#	print('removing reaction')
#	global reactId
##	print reactId.
#	for reaction in reactId():
#		if reaction not in numbers:
#			await reactId.clear(reaction)

@game.error
async def game_error(ctx, error):
	if isinstance(error, commands.MissingRequiredArgument):
		embed = discord.Embed(color = 0x607d8b, description = 'Please add in at least one option to be added on the poll. \
						\nExample: `!game Valorant Aram Urf`')
		await ctx.send(embed = embed)

@agents.error
async def agents_error(ctx, error):
	if isinstance(error, commands.MissingRequiredArgument):
		embed = discord.Embed(color = 0x607d8b, description = 'Please add in at least one user. \
						\nExample: `!agents @user1 @user2`')
		await ctx.send(embed = embed)



#Sends a question mark any message Zach sends
@client.event
async def on_message(message):
	if message.author.id == 219308580036280321: #Zach's ID
		await message.add_reaction('\u2753') #Question Mark Unicode

	await client.process_commands(message)

client.run('NzE1ODExMjE3OTAxNjE3MTUy.Xt8P_w.fLQqsCXf8W9-4WhbtbOXBHroBGY')

